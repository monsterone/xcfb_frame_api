# -*- coding: utf-8 -*-
# @Author   : Monster
# @File     : do_excel.py

from openpyxl import load_workbook
from tools.read_config import ReadConfig
from tools.project_path import *
class DoExcel():

    def __init__(self,file_name,sheet_name=None):
        self.file_name=file_name
        self.sheet_name=sheet_name

    def get_data(self):
        wb = load_workbook(self.file_name)
        # sheet = wb[self.sheet_name]
        #读取配置文件只
        mode=eval(ReadConfig.get_config(case_config_path,'MODE','mode'))

        test_data=[]
        for key in mode: #遍历存在这个字典里面的字典
            sheet = wb[key]
            if mode[key] == 'all':
                for i in range(2,sheet.max_row+1):
                    sub_data={}
                    sub_data["case_id"] = sheet.cell(i,1).value
                    sub_data["module"] = sheet.cell(i,2).value
                    sub_data["title"]=sheet.cell(i,3).value
                    sub_data["method"]=sheet.cell(i,4).value
                    sub_data["url"]=sheet.cell(i,5).value
                    sub_data["data"]=sheet.cell(i,6).value
                    sub_data["expected"]=sheet.cell(i,7).value
                    sub_data["sheet_name"]=key
                    test_data.append(sub_data)

            else:
                for case_id in mode[key]:
                    i = case_id+1  #行号
                    sub_data = {}
                    sub_data["case_id"] = sheet.cell(i, 1).value
                    sub_data["module"] = sheet.cell(i, 2).value
                    sub_data["title"] = sheet.cell(i, 3).value
                    sub_data["method"] = sheet.cell(i, 4).value
                    sub_data["url"] = sheet.cell(i, 5).value
                    sub_data["data"] = sheet.cell(i, 6).value
                    sub_data["expected"] = sheet.cell(i, 7).value
                    sub_data["sheet_name"] = key
                    test_data.append(sub_data)

        return test_data

    def write_back(self,i,value,TestResult):#专门写回数据
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]

        sheet.cell(i,8).value=value
        sheet.cell(i,9).value=TestResult
        wb.save(self.file_name)  #保存结果



if __name__ == '__main__':
    res=DoExcel(test_case_path).get_data()
    print(res)
    print(len(res))